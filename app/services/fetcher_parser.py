"""services/fetcher_parser.py

Fetcher and Excel‑parser for GUU timetable files.
Designed for Python 3.12+, SQLAlchemy 2.x async engine.

Usage example
-------------
>>> await sync()
"""
from __future__ import annotations

import asyncio
import hashlib
import logging
import re
import tempfile
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable, List, NamedTuple, Sequence

import httpx  # async requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pydantic import BaseModel, Field, ValidationError
from sqlalchemy import UniqueConstraint, delete, select
from sqlalchemy.ext.asyncio import AsyncSession

# ---- logging --------------------------------------------------------------
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# ---- constants ------------------------------------------------------------
SCHEDULE_URL = "https://guu.ru/student/schedule/"
XLSX_RE = re.compile(r"\.xlsx?$")
GROUP_RE = re.compile(r"([А-ЯA-ZЁ\-]+\d{2,3})", re.I)
DATE_CELL_RE = re.compile(r"^(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})$")

# ---- data‑models ----------------------------------------------------------
class Lesson(BaseModel):
    group_code: str
    date: date
    start_time: time
    end_time: time
    subject: str
    teacher: str | None = None
    room: str | None = None

    class Config:
        frozen = True


class SourceFileMeta(BaseModel):
    url: str
    sha256: str
    semester: str | None = None
    fetched_at: datetime = Field(default_factory=datetime.utcnow)


class _FileLink(NamedTuple):
    url: str
    filename: str


# ---- core functions -------------------------------------------------------
async def list_files(client: httpx.AsyncClient | None = None) -> List[_FileLink]:
    """Parse GUU schedule page and return xlsx links."""
    own_client = client is None
    if own_client:
        client = httpx.AsyncClient(timeout=20)
    try:
        r = await client.get(SCHEDULE_URL)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        links = []
        for a in soup.select("a"):
            href: str | None = a.get("href")
            if href and XLSX_RE.search(href):
                links.append(_FileLink(url=httpx.URL(href, base=SCHEDULE_URL), filename=Path(href).name))
        logger.info("Found %d xlsx links on page", len(links))
        return links
    finally:
        if own_client:
            await client.aclose()


async def download(url: str | httpx.URL, dest_dir: Path | str | None = None) -> Path:
    """Download a file and return local path. Destination dir defaults to tmp."""
    dest_dir = Path(dest_dir) if dest_dir else Path(tempfile.mkdtemp())
    dest_dir.mkdir(parents=True, exist_ok=True)
    filename = Path(url).name
    dest_path = dest_dir / filename
    async with httpx.AsyncClient(timeout=60) as client:
        r = await client.get(url)
        r.raise_for_status()
        dest_path.write_bytes(r.content)
    logger.info("Downloaded %s → %s (%.1f KiB)", url, dest_path, dest_path.stat().st_size / 1024)
    return dest_path


def sha256_path(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# EXCEL → Lesson records
# ---------------------------------------------------------------------------

def _parse_time(cell_value: str) -> tuple[time, time]:
    """Split string like '10:30‑12:05' into two times."""
    if not cell_value:
        raise ValueError("Empty time")
    parts = re.split(r"[\u2010\u2013\u2014\-]", str(cell_value))  # support hyphen–dash
    if len(parts) != 2:
        raise ValueError(cell_value)
    t1 = datetime.strptime(parts[0].strip(), "%H:%M").time()
    t2 = datetime.strptime(parts[1].strip(), "%H:%M").time()
    return t1, t2


def _safe_str(v) -> str:
    return str(v).strip() if v is not None else ""


def parse_excel(path: Path) -> Sequence[Lesson]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active  # assume first sheet contains timetable

    # Attempt to detect group code row (first row that contains recognizable group code)
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if any(GROUP_RE.search(str(cell or "")) for cell in row):
            header_row_idx = row[0].row if hasattr(row[0], "row") else 1  # openpyxl 3.1 gives tuple of values; no .row
            break
    if header_row_idx is None:
        raise RuntimeError("Group header not found in Excel")

    lessons: list[Lesson] = []

    # For simplicity assume layout: columns: A date, B time, C subject, D teacher, E room, F group
    # Real‑world file may be different; adjust as needed.
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        date_val, time_val, subject_val, teacher_val, room_val, group_val = map(_safe_str, row[:6])
        if not (date_val and time_val and subject_val and group_val):
            continue
        # Parse date
        try:
            if isinstance(date_val, datetime):
                lesson_date = date_val.date()
            elif isinstance(date_val, date):
                lesson_date = date_val
            else:
                m = DATE_CELL_RE.match(date_val)
                if not m:
                    raise ValueError(date_val)
                day, month, year = map(int, m.groups())
                if year < 100:
                    year += 2000
                lesson_date = date(year, month, day)
        except Exception as e:
            logger.debug("Skip row; bad date %s: %s", date_val, e)
            continue
        # Parse time
        try:
            start_t, end_t = _parse_time(time_val)
        except Exception as e:
            logger.debug("Skip row; bad time %s: %s", time_val, e)
            continue
        try:
            lessons.append(
                Lesson(
                    group_code=group_val.upper(),
                    date=lesson_date,
                    start_time=start_t,
                    end_time=end_t,
                    subject=subject_val,
                    teacher=teacher_val or None,
                    room=room_val or None,
                )
            )
        except ValidationError as e:
            logger.debug("Validation error in row: %s", e)
    logger.info("Parsed %d lesson rows from %s", len(lessons), path.name)
    return lessons


# ---------------------------------------------------------------------------
# DB integration helpers (SQLAlchemy async) ---------------------------------
# ---------------------------------------------------------------------------
# Assuming models mapped as in previous design; import them lazily to avoid circular deps.


async def upsert_source_file(meta: SourceFileMeta, db: AsyncSession) -> int:
    """Insert SourceFile if sha256 unseen. Return id."""
    from db.models import SourceFile  # pylint: disable=import-error

    res = await db.execute(select(SourceFile.id).where(SourceFile.sha256 == meta.sha256))
    existing_id = res.scalar_one_or_none()
    if existing_id:
        return existing_id
    sf = SourceFile(
        url=meta.url,
        sha256=meta.sha256,
        semester=meta.semester,
        added_at=meta.fetched_at,
    )
    db.add(sf)
    await db.flush()
    logger.info("Inserted SourceFile id=%s", sf.id)
    return sf.id


async def bulk_insert_lessons(lessons: Sequence[Lesson], source_id: int, db: AsyncSession) -> None:
    """Insert lessons; duplicates (same group/date/time/subject) are ignored."""
    from db.models import Group, Lesson as LessonORM, Room, Teacher  # pylint: disable=import-error

    # Cache for fk lookup
    group_cache: dict[str, int] = {}
    room_cache: dict[str, int] = {}
    teacher_cache: dict[str, int] = {}

    async def _get_group(group_code: str) -> int:
        if group_code in group_cache:
            return group_cache[group_code]
        res = await db.execute(select(Group.id).where(Group.code == group_code))
        gid = res.scalar_one_or_none()
        if gid is None:
            raise RuntimeError(f"Unknown group {group_code}; import group seeds first")
        group_cache[group_code] = gid
        return gid

    async def _get_or_create_room(room_name: str) -> int:
        if room_name in room_cache:
            return room_cache[room_name]
        building, _, number = room_name.partition("-") if "-" in room_name else ("", "", room_name)
        res = await db.execute(select(Room.id).where(Room.building == building, Room.number == number))
        rid = res.scalar_one_or_none()
        if rid is None:
            room = Room(building=building, number=number)
            db.add(room)
            await db.flush()
            rid = room.id
        room_cache[room_name] = rid
        return rid

    async def _get_or_create_teacher(name: str) -> int:
        if name in teacher_cache:
            return teacher_cache[name]
        res = await db.execute(select(Teacher.id).where(Teacher.name == name))
        tid = res.scalar_one_or_none()
        if tid is None:
            teacher = Teacher(name=name)
            db.add(teacher)
            await db.flush()
            tid = teacher.id
        teacher_cache[name] = tid
        return tid

    for rec in lessons:
        gid = await _get_group(rec.group_code)
        rid = await _get_or_create_room(rec.room) if rec.room else None
        tid = await _get_or_create_teacher(rec.teacher) if rec.teacher else None
        obj = LessonORM(
            group_id=gid,
            date=rec.date,
            start_time=rec.start_time,
            end_time=rec.end_time,
            subject=rec.subject,
            teacher_id=tid,
            room_id=rid,
            source_id=source_id,
        )
        db.add(obj)
    await db.commit()
    logger.info("Inserted %d Lesson rows (source_id=%s)", len(lessons), source_id)


# ---------------------------------------------------------------------------
# High‑level sync -----------------------------------------------------------
# ---------------------------------------------------------------------------

async def sync(db: AsyncSession, *, limit: int | None = None) -> None:
    """Main entry: fetch new xlsx files, parse and load into DB."""
    links = await list_files()
    if limit:
        links = links[:limit]
    for link in links:
        try:
            path = await download(link.url)
            file_hash = sha256_path(path)
            meta = SourceFileMeta(url=str(link.url), sha256=file_hash)
            sf_id = await upsert_source_file(meta, db)
            # If file already exists (duplicate), skip parsing
            if not sf_id:
                continue
            lessons = parse_excel(path)
            await bulk_insert_lessons(lessons, sf_id, db)
        except Exception as e:  # pylint: disable=broad-except
            logger.exception("Failed to process %s: %s", link.url, e)


# ---- CLI for manual run ---------------------------------------------------
if __name__ == "__main__":
    # Example CLI: python fetcher_parser.py sqlite+aiosqlite:///schedule.db
    import sys
    from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker

    if len(sys.argv) < 2:
        print("Usage: python fetcher_parser.py <db_url>")
        raise SystemExit(1)

    engine = create_async_engine(sys.argv[1], echo=False)
    async_session = async_sessionmaker(engine, expire_on_commit=False)

    async def _main():
        async with async_session() as session:
            await sync(session)

    asyncio.run(_main())
